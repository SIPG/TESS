#!/usr/bin/python
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import openpyxl
import sys
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border,colors
from openpyxl.styles import Alignment
from sys import argv
import urllib, json
from optparse import OptionParser
from openpyxl.styles.borders import Border, Side

arguments=sys.argv
RFA_link=arguments[1]
Cluster_ID=arguments[2]
wb = Workbook()
wb=openpyxl.load_workbook("RFA_sheet.xlsx")
sheets=wb.get_sheet_names()

def modify(url_env,sheet,drop):
	failed_list,jira_list,team_list,comment_list=[],[],[],[]
        url=url_env.replace("environment","xunit")
        response = urllib.urlopen(url)
        data = json.loads(response.read())
        suites_list = data["testSuites"]
        suites_in_allure=[]
        total_cases,passed_cases=0,0
        for item in suites_list:
                suites_in_allure.append( item["name"])
       		suites_in_allure=sorted(suites_in_allure)
	        total_cases=total_cases+int(item["statistic"]["total"])
                passed_cases=passed_cases+int(item["statistic"]["passed"])
                if item["statistic"]["passed"] != item["statistic"]["total"]:
                        failed_list.append(item["name"])
	count=0
	for name in sheets:
                sheet_name="RFA250_"+drop
		if sheet_name == name:
			sheet_num=sheets.index(sheet_name)
			count+=1
	if count==0:
		print "Adding new Sheet for ", drop
		suites_in_allure.insert(0,"Test Suites")
		wb.create_sheet(sheet_name,1)
		ws2 = wb.get_sheet_by_name(name=sheet_name)
		c=ws2['B2']
		ws2.freeze_panes = c
		for suite in range(0,len(suites_in_allure)):
			cell_A="A"+str(suite+1)
			ws2[cell_A].value=suites_in_allure[suite]
	else:
		print "Modifying in RFA250 sheet",drop," for ",ISO
	        ws2 = wb.get_sheet_by_name(name = sheets[sheet_num])
        c=ws2['B1']
        ws2.freeze_panes = c
        global suites_in_sheet
        suites_in_sheet=[]
        column_count=ws2.max_column
        for i in range (2,ws2.max_row+1):
                cell_A_sheet="A"+str(i)
                suites_in_sheet.append(ws2[cell_A_sheet].value)
        blackFill = PatternFill(start_color='FF000000',end_color='FF000000',fill_type='solid')
        redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
        whiteFill = PatternFill(start_color='FFFFFFFF',end_color='FFFFFFFF',fill_type='solid')
	orangeFill = PatternFill(start_color='ffa500',end_color='ffa500',fill_type='solid')
	thick_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        percent=(float(passed_cases)/float(total_cases))*100
        cell=str(chr(column_count/26+64))+str(chr(column_count%26+65))
        cell=cell.replace("@","")
        ws2[cell+"1"]=ISO+"("+server+")" + "%.2f" % percent+"%"
        cal=Font(name='Calibri',size=11)
        ws2[cell+"1"].font=cal
        ws2[cell+"1"].hyperlink = (url[:-15]+"#/xunit")
        for j in failed_list:
                for i in range (2,ws2.max_row+1):
                        cell_A="A"+str(i)
			ws2[cell_A].fill = whiteFill
                        if ws2[cell_A].value==j:
                                cell_red=cell+cell_A[1:]
                                ws2[cell_red].fill = redFill
                                print j
	for suite in suites_in_sheet:
                         if suite not in suites_in_allure and suite !=None:
                                remove_cell=cell+str(suites_in_sheet.index(suite)+2)
                                ws2[remove_cell].fill=orangeFill
        cell=str(chr((column_count)/26+64))+str(chr((column_count)%26+65))
        cell=cell.replace("@","")
        cell=cell+"1"
        ws2[cell].font=Font(bold=True, size=10)
	for row1 in range(1,ws2.max_row):
		for col1 in range(1,ws2.max_column):
			ws2.cell(row=row1, column=col1).border = thick_border
	ws2.column_dimensions["A"].width = 55
	col1= ws2.max_column-1
	col=str(chr((col1)/26+64))+str(chr((col1)%26+65))
	if "@"in col:
		col=col.replace("@","")
        ws2.column_dimensions[col].width = 20
reruns=[]
if "xunit" in RFA_link:
        url_env = RFA_link[:-8]+"/data/environment.json"
elif "#" in RFA_link:
        url_env = RFA_link[:-3]+"/data/environment.json"
else:
        url_env = RFA_link+"data/environment.json"
response_env = urllib.urlopen(url_env)
data_env = json.loads(response_env.read())
info_list= data_env["parameter"]
server=info_list[0]["value"]
ISO=info_list[9]["value"]
drop=info_list[3]["value"]
print "**********************RFA 250************************"
modify(url_env,1,drop)
sys.stdout.flush()

wb.save('RFA_sheet.xlsx')
wb.save('RFA_sheet_copy.xlsx')


