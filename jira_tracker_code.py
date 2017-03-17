from openpyxl.styles import Font
import openpyxl
import collections
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
from lxml import html
import requests
import urllib, urllib2, cookielib
import re
import time
from datetime import datetime
username = 'xsuhchi'
password = 'suhrud@03'
page_source=""
wb = openpyxl.Workbook()
wb=openpyxl.load_workbook("JIRA_tracker.xlsx")
sheets=wb.get_sheet_names()
ws = wb.get_sheet_by_name(name = sheets[0])
ws["A1"].value="Name"
italic24Font = Font(size=12, bold=True)
ws['A1'].font = italic24Font
ws["B1"].value="JIRA"
ws['B1'].font = italic24Font
ws["C1"].value="Comments"
ws['C1'].font = italic24Font
TORF_list,RTD_list,CIP_list,CIS_list,NSS_list=[],[],[],[],[]
cj = cookielib.CookieJar()
opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cj))
login_data = urllib.urlencode({'os_username' : username, 'os_password' : password})
opener.open('https://confluence-nam.lmera.ericsson.se/login.action?os_destination=%2Fpages%2Fviewpage.action%3FspaceKey%3DEM%26title%3DOffshore%2BJira%2BTracker&permissionViolation=true', login_data)
resp = opener.open('https://confluence-nam.lmera.ericsson.se/pages/viewpage.action?spaceKey=EM&title=Offshore+Jira+Tracker')
page_source=resp.read()
def findOccurences(s, ch):
	return [m.start() for m in re.finditer(ch,s )]
TORF_in_source=findOccurences(page_source, "TORF")
RTD_in_source=findOccurences(page_source, "RTD")
CIP_in_source=findOccurences(page_source, "CIP")
CIS_in_source=findOccurences(page_source, "CIS")
NSS_in_source=findOccurences(page_source, "NSS")
tickets_in_source=TORF_in_source+RTD_in_source+CIP_in_source+CIS_in_source+NSS_in_source
print len(tickets_in_source)
tickets_in_source=sorted(tickets_in_source)
print len(tickets_in_source)
ticket_list=[]
for i in tickets_in_source:
	if page_source[i:i+4]=="TORF":
		ticket_list.append(page_source[i:i+11])
	elif page_source[i:i+3]=="RTD":
		if page_source[i+6] in ["1","2","3","4","5","6","7","8","9","0"]:
        	        ticket_list.append(page_source[i:i+7])
        	else:
        	        ticket_list.append(page_source[i:i+6]) 
	elif page_source[i:i+3]=="NSS":
	                ticket_list.append(page_source[i:i+8])
	else:
		ticket_list.append(page_source[i:i+9])
l2 = []
s = set()
for i in ticket_list:
   if not i in s:
       l2.append(i)
       s.add(i)
opener.open('https://jira-nam.lmera.ericsson.se/login.jsp?permissionViolation=true&os_destination=%2Fbrowse%2FTORF-136714&page_caps=&user_role=', login_data)
tree = html.fromstring(page_source)
prices = tree.xpath('//td[@class="jira-macro-table-underline-pdfexport"]/text()')
j=2
for i in range(0,len(prices)):
	prices[i]=prices[i].strip('\n').strip(' ').rstrip('\n').rstrip(' ')
date_created_list,date_modified_list,assigned_to_list,created_by_list=[],[],[],[]
for i in range(6,len(prices),16):
        prices[i]=re.sub(',', '', prices[i])
        date_object = datetime.strptime(prices[i], '%b %d %Y')
	date_created_list.append(date_object)
for i in range(7,len(prices),16):
	prices[i]=re.sub(',', '', prices[i])
	date_object = datetime.strptime(prices[i], '%b %d %Y')
        date_modified_list.append(date_object)
for i in range(10,len(prices),16):
        assigned_to_list.append(prices[i])
        cell_A="A"+str(j)
        ws[cell_A].value=prices[i]
        j=j+1
j=2
for i in range(9,len(prices),16):
        created_by_list.append(prices[i])
for i in range(0,len(l2)):
        delta=datetime.now()-date_modified_list[i]
	cell_C="C"+str(j)
	if created_by_list[i] =="Unassigned":
		ws[cell_C]="Unassigned"
	elif delta.days<10:
		ws[cell_C]="Updated"
	else:
		ws[cell_C]="No update since "+str(date_modified_list[j-2].date())
	j=j+1
j=2
for i in l2:
	cell_B="B"+str(j)
        ws[cell_B].value=i
        j=j+1
print len(l2)
wb.save("JIRA_tracker.xlsx")
#MTG
