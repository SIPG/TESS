#!/usr/bin/python
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import openpyxl
import sys
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border,colors
from openpyxl.styles import Alignment
from sys import argv
import urllib, json
from optparse import OptionParser
arguments=sys.argv
print arguments[1:]
wb = Workbook()
wb=openpyxl.load_workbook('RFA_sheet.xlsx')
sheets=wb.get_sheet_names()
suites_in_sheet=[]
global suites_in_sheet
ws2 = wb.get_sheet_by_name(name = sheets[1])
column_count=ws2.max_column
cell=str(chr((column_count-3)/26+64))+str(chr((column_count-3)%26+64))
for i in range (2,ws2.max_row+1):
	cell_A_sheet="A"+str(i)
	suites_in_sheet.append(ws2[cell_A_sheet].value)
whiteFill = PatternFill(start_color='FFFFFFFF',end_color='FFFFFFFF',fill_type='solid')
def rerun(suite,link,ISO):
	for i in range (ws2.max_column,0,-1):
	        cell=str(chr((i)/26+64))+str(chr((i)%26+64))
		cell=cell.replace("@","")
		if ISO in ws2[cell+"1"].value:
        		for item in range(0,len(suites_in_sheet)):
		                if suites_in_sheet[item]==suite:
                		        ws2[cell+str(item+2)].value="Passed in rerun"
                        		ws2[cell+str(item+2)].hyperlink = (link)
					ws2[cell+str(item+2)].fill=whiteFill
					print cell+str(item+2)
					print suites_in_sheet[item]
			break
def group(rerun_links):
               for item in rerun_links:
			env_link=str(item[:-3])+"/data/environment.json"
                        response_env = urllib.urlopen(env_link)
                        data_env = json.loads(response_env.read())
			info_list= data_env["parameter"]
			ISO=info_list[9]["value"]
                        link=str(item[:-3])+"/data/xunit.json"
                        response = urllib.urlopen(link)
                        data = json.loads(response.read())
                        suite= data["testSuites"][0]["title"]
			print suite
                        rerun(suite,item,ISO)
               wb.save('RFA_sheet.xlsx')
               sys.exit(0)
rerun_links=arguments[1:]
group(rerun_links)

